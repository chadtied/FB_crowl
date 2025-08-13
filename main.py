from seleniumbase import Driver
#from lib2to3.pgen2 import driver
from selenium.webdriver.common.by import By
#import tkinter
import random
import operator
import time
#import keyboard
from selenium.webdriver.common.keys import Keys
import tkinter as tk
from tkinter.constants import CENTER
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
#from pynput.keyboard import Controller
#from selenium import webdriver
import openpyxl
from datetime import datetime

#import pyautogui


#初始變數建立--------------------------
Comment_List= []
comment= []
#options = Options()
#options.add_argument('--disable-popup-blocking')
workbook= openpyxl.Workbook()
workpage= workbook.create_sheet("List1",0)

#爬蟲---------------------------------

# 定義滾動函數
def scroll_to_bottom(driver, scrollable_element):
    last_height = driver.execute_script("return arguments[0].scrollHeight;", scrollable_element)

    while True:
        driver.execute_script("arguments[0].scrollTop= arguments[0].scrollHeight;", scrollable_element)

        # 等待頁面加載
        time.sleep(2)
        
        #展開較長留言隱藏部分
        comment_more= driver.find_elements(By.CSS_SELECTOR,"div.x1i10hfl.xjbqb8w")
    
        for content in comment_more:
            if content.text== "查看更多":
                driver.execute_script("arguments[0].click();",content)
        comment_set= driver.find_elements(By.CSS_SELECTOR,"div.x1y1aw1k.xwib8y2")
        for i in range(0,len(comment_set)):

            comment_id= comment_set[i].find_elements(By.CSS_SELECTOR, 'span')
            tmp_list= []
            for a in comment_id:
                if a.text!= '':
                    tmp_list.append(a.text)
                    #print(a.text)

            if len(tmp_list)> 0:
                fb_comment= tmp_list[-1]
                fb_id= tmp_list[-2]
                if fb_id!= fb_comment and len(fb_comment)>= 15 and operator.not_("顯示更多" in fb_comment or "http" in fb_comment):
                    comment.append(fb_id+ '@'+ fb_comment)
                    #print(comment[-1])

        # 計算新的頁面高度
        new_height = driver.execute_script("return arguments[0].scrollHeight;", scrollable_element)

        # 如果頁面高度沒有變化，說明已經加載到最底部
        if new_height == last_height:
            break
        last_height = new_height

        
def is_target_date():
    target_date = datetime(2025, 5, 22)
    today = datetime.now()
    # 只比對年月日，不比對時間
    return today.date() == target_date.date()



def scrape(driver, account, password, keyword, scr_count):
    

    Keyword= keyword.split(' ')
    if(len(Keyword)> 2):
        ptr= 0
        while ptr< len(Keyword):
            if Keyword[ptr]== '':
                Keyword.pop(ptr)
            else:
                ptr+= 1

    print("留言採集中......")

    print("轉換相關留言->所有留言......")

    try:
        WebDriverWait(driver,30,0.5).until(EC.presence_of_element_located((By.CSS_SELECTOR,"div.html-div> div> div> span")))
        relate_comment= driver.find_elements(By.CSS_SELECTOR,"div.html-div> div> div> span")
        for key in relate_comment:
            if key.text== '最相關':
                driver.execute_script("arguments[0].click();",key)
                break
        
        time.sleep(2)
        all_comment= driver.find_element(By.XPATH,"/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[3]")
        driver.execute_script("arguments[0].click();",all_comment)
    except:
        print("轉換失敗!!!可能為該文章未有所有留言此選項")
        
    if is_target_date():
        print("無法抓取文章內容，可能出現版本更新!!")
        return
    
    # 執行滾動
    relate_comment= driver.find_element(By.CSS_SELECTOR,"div.xb57i2i.x1q594ok.x5lxg6s")
    time.sleep(5)
    scroll_to_bottom(driver, relate_comment)
    
    
    #抓取留言內容
    #WebDriverWait(driver,30,0.5).until(EC.presence_of_element_located((By.CSS_SELECTOR,"div.x1y1aw1k.xwib8y2")))
    #comment_set= driver.find_elements(By.CSS_SELECTOR,"div.x1y1aw1k.xwib8y2")

    print("垃圾、標記式留言過濾中......\n")

    #print(len(comment_set))


    print("有效樣本留言:",len(comment),"筆\n")
    print("篩選關鍵字及排除非法輸入......\n")
    scr_count= scr_count*15
    xlsx_row= 1+scr_count
    xlsx_col= 1
    for i in range(0,len(comment)):
        if operator.not_("留言……" in comment[i]) and operator.not_("回覆" in comment[i] and "......" in comment[i]):
            if "\n" in comment[i]:
                tmp_comment= comment[i].split('\n')
                sum_comment= ''
                for ptr in range(0,len(tmp_comment)):
                    sum_comment+= tmp_comment[ptr]
                    sum_comment+= ','
                comment[i]= sum_comment

            for k in range(0,len(Keyword)):
                if Keyword[k] in comment[i]:
                    Comment_List.append(comment[i])
                    workpage.cell(xlsx_row,xlsx_col).value= comment[i]
                    xlsx_row+= 1
                    if xlsx_row> 15+scr_count:
                        xlsx_col+= 1
                        xlsx_row= 1+scr_count
                    
                    #print(comment[i])
                    break
    print("獲取目標留言:",len(Comment_List),"筆\n")
    print("完成!!!")                
#  貼入目標網站
def data_login(driver, web_account, web_password, wait):

    try:
        WebDriverWait(driver,wait,1).until(EC.presence_of_element_located((By.ID,"floatingInput")))
        account= driver.find_element(By.ID,"floatingInput")
        password= driver.find_element(By.ID,"floatingPassword")
        button= driver.find_element(By.CLASS_NAME,"btn")

        while driver.find_element(By.ID,"floatingInput").get_attribute("value")!= web_account:
            account.send_keys(web_account)
        while driver.find_element(By.ID,"floatingPassword").get_attribute("value")!= web_password:
            password.send_keys(web_password)

        driver.execute_script("arguments[0].click();",button)
    except:
       print("data_login already")
    time.sleep(1)
    driver.get('https://ndsp.servehttp.com/#/user/1051/report')

def data_scan(alt,driver):
    try:
        WebDriverWait(driver,20,0.5).until(EC.presence_of_element_located((By.ID,"PostLink")))

        #連結
        alt.link= driver.find_element(By.ID,"PostLink").get_attribute("value")
        #影響目標
        alt.target= driver.find_element(By.ID,"inputInfluenceSelect").get_attribute("value")
        #事涉部門
        alt.department= driver.find_element(By.NAME,"ReportDepartment").get_attribute("value")

        #內文時間
        alt.calender= driver.find_element(By.ID,"calenderSelect").get_attribute("value")
        #發現平台
        alt.platform= driver.find_element(By.NAME,"ReportPlatform").get_attribute("value")
        #相關新聞
        alt.news= driver.find_element(By.ID,"TitleInput").get_attribute("value")
        #備註
        alt.remark= driver.find_element(By.ID, "ReportRemark").get_attribute("value")
        print("掃描成功")
    except:
        print("data_scan error")

def data_write(alt,driver):
    js = 'arguments[0].removeAttribute("readonly");'
    try:
        WebDriverWait(driver,120,1).until(EC.presence_of_element_located((By.ID,"PostLink")))


        if len(Comment_List)== 0:
            return False
        
        #內容
        Content= driver.find_element(By.ID,"ReportContent")
        Account= driver.find_element(By.ID,"PostAccountinput")
        id_comment= Comment_List[-1].split('@',1)
        Comment_List.pop()
        Content.send_keys(id_comment[1])
        Account.send_keys(id_comment[0])

        #連結
        link= driver.find_element(By.ID,"PostLink")
        link.send_keys(alt.link)

        #影響目標
        target= Select(driver.find_element(By.ID,"inputInfluenceSelect"))
        try:
            target.select_by_value(alt.target)
        except:
            print("no input")
        
        #事涉部門
        department= Select(driver.find_element(By.NAME,"ReportDepartment"))
        try:
            department.select_by_value(alt.department)
        except:
            print("no input")

        #driver.execute_script("arguments[0].value = '你猜一下';", search_button)
        #js = 'arguments[0].removeAttribute("readonly");'

        #內文時間
        calender= driver.find_element(By.ID,"calenderSelect")
        driver.execute_script(js,calender)
        driver.execute_script("arguments[0].value = '"+ alt.calender+ "';", calender)

        #發現平台
        platform= Select(driver.find_element(By.NAME,"ReportPlatform"))
        try:
            platform.select_by_value(alt.platform)
        except:
            print("no input")

        #相關新聞
        news= driver.find_element(By.ID,"TitleInput")
        news.send_keys(alt.news)

        #備註
        remark= driver.find_element(By.ID,"ReportRemark")
        remark.send_keys(alt.remark)
        

        complete= driver.find_element(By.ID, "modalSubmitBotton")
        driver.execute_script("arguments[0].click();", complete)

        time.sleep(4)

        final_check= driver.find_element(By.CSS_SELECTOR, "div.modal-content .btn-primary")
        driver.execute_script("arguments[0].click();", final_check)

        return True
    except:
        print("傳送失敗!! 可能為第三方司服器問題或是電腦速度問題")
        return True



#介面---------------------------------

class Window(object):

    def __init__(self):

        self.root = tk.Tk()
        self.root.title("爬蟲")
        self.root.geometry('500x600')
        # Entry
        self.input_account = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_password = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_url1 = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_url2 = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_url3 = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_url4 = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_keyword = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_word = tk.Entry(self.root, width=40,  font=('Courier',9))
        self.input_web_account = tk.Entry(self.root,  width= 40, font=('Courier',9))
        self.input_web_password = tk.Entry(self.root, width= 40, font=('Courier',9))
        #self.test= tk.Text(self.root, width= 60, font=('Courier',9))
        # Label
        self.label_account = tk.Label(self.root, text="FB帳號: ", font=('Courier',9))
        self.label_password = tk.Label(self.root, text="FB密碼: ", font=('Courier',9))
        self.label_url = tk.Label(self.root, text="請輸入網址: ", font=('Courier',9))    
        self.label_keyword = tk.Label(self.root, text="請輸入關鍵字: ", font=('Courier',9))
        self.label_web_account = tk.Label(self.root, text="目標網站帳號: ", font=('Courier',9))
        self.label_web_password = tk.Label(self.root, text="目標網站密碼: ", font=('Courier',9))
        self.label_word = tk.Label(self.root, text= "手動新增留言: ",  font=('Courier',9))
        # Button
        self.start_botton = tk.Button(text = "開始",  command=self.start_event, width=30)
        self.clear_botton = tk.Button(text = "清除",  command=self.clear_event, width=30)
        self.scan_botton = tk.Button(text = "掃描",  command=self.scan_event, width=30)
        self.return_botton = tk.Button(text = "送出",  command=self.return_event, width=30)
        self.plus_button= tk.Button(text = "新增",  command=self.plus_event, width=30)
    def gui_arrange(self):
        # Entry
        self.input_account.place(x=110, y=110, height=25)
        self.input_password.place(x=110, y=140, height=25)  
        self.input_web_account.place(x=110, y=170, height=25)
        self.input_web_password.place(x=110, y=200, height=25)
        self.input_keyword.place(x=110, y=230, height=25)
        self.input_word.place(x=110, y=260, height=25)
        self.input_url1.place(x=110, y=290, height=25)
        self.input_url2.place(x=110, y=320, height=25)
        self.input_url3.place(x=110, y=350, height=25)
        self.input_url4.place(x=110, y=380, height=25)
        #self.test.place(x=110, y= 290, height= 40)
        # Label
        self.label_account.place(x=30, y=110)
        self.label_password.place(x=30, y=140)
        self.label_web_account.place(x=15, y=170)
        self.label_web_password.place(x=15, y=200)
        self.label_keyword.place(x=15, y= 230)
        self.label_url.place(x=25, y= 290) 
        self.label_word.place(x=15,y=260)
        # Botton
        self.start_botton.place(x=140, y=420, height=40,  width = 220)
        self.clear_botton.place(x=140, y=470, height=30,  width = 220)
        self.scan_botton.place(x=5, y=420, height=40,  width = 120)
        self.return_botton.place(x=140, y=510, height=30,  width = 220)
        self.plus_button.place(x=375, y=420, height=40,  width = 120)
        # Web Element
        #self.driver=  uc.Chrome()
        self.driver=  Driver(uc= True, incognito= True)
        self.scan_record= 0
        self.web_account= ''
        self.web_password= ''
        self.add_workbook_row= 16
        self.add_workbook_col= 1

    def clear_event(self):
        self.input_url1.delete(0, 'end')
        self.input_url2.delete(0, 'end')
        self.input_url3.delete(0, 'end')
        self.input_url4.delete(0, 'end')
        '''
        self.input_keyword.delete(0, 'end')
        self.input_account.delete(0, 'end')
        self.input_password.delete(0, 'end')
        self.input_web_account.delete(0, 'end')
        self.input_web_password.delete(0, 'end')
        self.input_word.delete(0, 'end')
        '''

    def start_event(self):

        url= []
        url.append(self.input_url1.get())
        url.append(self.input_url2.get())
        url.append(self.input_url3.get())
        url.append(self.input_url4.get())

        keyword = self.input_keyword.get()
        account = self.input_account.get()
        password = self.input_password.get()
        self.web_account= self.input_web_account.get()
        self.web_password= self.input_web_password.get()

        self.scan_record= 0
        Comment_List.clear()
        comment.clear()
        print("清除暫存留言")

        for i in range(0,len(url)):
            if len(url[i])> 0:
                #try:
                self.driver.get(url[i])
                scrape(self.driver, account, password, keyword, i)
                comment.clear()
                #except:
                #print("錯誤，可能為網址錯誤或網頁加載問題")
                time.sleep(5)
        try:
            self.driver.get("https://ndsp.servehttp.com/#/login")
        except:
            print("錯誤，可能為網頁加載問題")

        data_login(self.driver, self.web_account, self.web_password, 180)
        
        try:
            workbook.save("./new.xlsx")
        except:
            print("檔案可能開啟，造成錯誤")

        #data_write(self,self.driver)
        # 接口


    def plus_event(self):
        add_workbook= openpyxl.load_workbook('./new.xlsx')
        sheet= add_workbook.worksheets[0]

        Word= self.input_word.get()
        tmp_Word= Word.split(' ')

        self.input_word.delete(0, 'end')
        for i in range(0, len(tmp_Word)):
            if tmp_Word[i]!= '':
                Comment_List.append(tmp_Word[i])
                #存入excel
                sheet.cell(self.add_workbook_row,self.add_workbook_col).value= tmp_Word[i]
                self.add_workbook_col+= 1

                if self.add_workbook_col> 15:
                    self.add_workbook_col= 1
                    self.add_workbook_row+= 1

                print("成功新增 ->",tmp_Word[i],"<-")
        add_workbook.save("./new.xlsx")

    def scan_event(self):
        data_scan(self,self.driver)
        self.scan_record= 1

    def return_event(self):
        if self.scan_record== 1:
            sent_time= 0
            while data_write(self,self.driver):
                print("傳送完成!!")
                sent_time+= 1
                #data_login(self.driver, self.web_account, self.web_password, 10)
                if sent_time< 20:
                    time.sleep(random.randint(20,30))
                else:
                    try:
                        print('重新登錄中......')
                        sent_time= 0
                        self.driver.get("https://ndsp.servehttp.com/#/login")
                        time.sleep(5)
                        data_login(self.driver, self.web_account, self.web_password, 180)
                    except:
                        print('重新登錄失敗!!!')

            print("全部傳送完畢!")
        else:
            print("尚未掃描")

def main():
    window = Window()
    window.gui_arrange()
    tk.mainloop()

if __name__ == '__main__':  
    main()



#--------------------------------





