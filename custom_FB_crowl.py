import undetected_chromedriver as uc
#from lib2to3.pgen2 import driver
from selenium.webdriver.common.by import By
import tkinter
import random
import operator
import time
import keyboard
from selenium.webdriver.common.keys import Keys
import tkinter as tk
from tkinter.constants import CENTER
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from pynput.keyboard import Controller
from selenium import webdriver
import openpyxl
import pyautogui


#初始變數建立--------------------------
Comment_List= []
comment= []
options = webdriver.ChromeOptions()
prefs= {
    'profile.default_content_setting_values' :
        {
        'notifications' : 2
        }
}
options.add_experimental_option('prefs',prefs)
workbook= openpyxl.Workbook()
workpage= workbook.create_sheet("List1",0)

#爬蟲---------------------------------


def scrape(driver, account, password, keyword):

    Keyword= keyword.split(' ')

    #登入介面
    try:
        WebDriverWait(driver,30,0.5).until(EC.presence_of_element_located((By.ID,"email")))

        while driver.find_element(By.ID,"email").get_attribute("value")!= account:
            Email= driver.find_element(By.ID,"email")
            Email.send_keys(account)
        while driver.find_element(By.ID,"pass").get_attribute("value")!= password:
            Password= driver.find_element(By.ID,"pass")
            Password.send_keys(password)

        time.sleep(2)
        Login= driver.find_element(By.ID,"loginbutton")
        driver.execute_script("arguments[0].click();", Login)
    except:
        print("FB already login")

    print("留言採集中......")

    '''
    # 抓取"相關留言"物件，並點選
    ActionChains(driver).move_by_offset(600,500).click().perform()
    WebDriverWait(driver,30,0.5).until(EC.presence_of_element_located((By.CSS_SELECTOR,"div.x1jx94hy.x12nagc")))
    form= driver.find_elements(By.CSS_SELECTOR,"div.x1jx94hy.x12nagc")

    span= form[-1].find_elements(By.CSS_SELECTOR,"span.x193iq5w")
    #form_click= driver.find_element(By.CSS_SELECTOR,".x78zum5> span")
    driver.execute_script("arguments[0].click();", span[0])


    #點選"所有留言"
    WebDriverWait(driver,30,0.5).until(EC.presence_of_element_located((By.CSS_SELECTOR,"div.xb57i2i.x1q594ok.x5lxg6s.x6ikm8r.x1ja2u2z > div")))
    List= driver.find_elements(By.CSS_SELECTOR,"div.xb57i2i.x1q594ok.x5lxg6s.x6ikm8r.x1ja2u2z > div")
    for i in range(0,len(List)):
       if "留言" in List[i].text:
           find_all= List[i].find_elements(By.CSS_SELECTOR,"span")
    
        
    driver.execute_script("arguments[0].click();",find_all[-1])
    '''

    #抓取第一則貼文
    #time.sleep(3)
    try:
        #ActionChains(driver).move_by_offset(600,500).click().perform()
        WebDriverWait(driver,30,0.5).until(EC.presence_of_element_located((By.CLASS_NAME,"x78zum5.x1n2onr6.xh8yej3")))
        article_num= 0
        pretend= 1
        count= 10000

        while True:
            count-= 1
            new_comment= driver.find_elements(By.CLASS_NAME,"x78zum5.x1n2onr6.xh8yej3")
            for i in range(0,len(new_comment)):
                if "留言" in new_comment[i].text:
                    article_num= i
                    pretend= 0
                    break
            if pretend== 0:
                break
            elif count< 0:
                print("can't not get data, maybe wrong type!")
                break
        
        Web= new_comment[article_num]
    except:
        print("first object catch fail")

   #print(new_comment[article_num].text)


    #展開留言內容
    '''
    flag= 0
    start= 0
    WebDriverWait(driver,30,0.5).until(EC.presence_of_element_located((By.CLASS_NAME,"x1i10hfl.xjbqb8w")))
    while flag!= 1:
        flag= 1
        time.sleep(5)
        more= Web.find_elements(By.CLASS_NAME,"x1i10hfl.xjbqb8w")
        #print(len(more))
        for i in range(start,len(more)):
            #print(more[i].text)
            if "檢視另" in more[i].text or (operator.not_("隱藏" in more[i].text) and "則回覆" in more[i].text)or "顯示更多" in more[i].text or "查看更多" in more[i].text:
                driver.execute_script("arguments[0].click();",more[i])
                time.sleep(3)
                flag= 0
                #print("yes")
        start= len(more)-1
        #more.clear
    '''
    #抓取相關留言
    flag= 0
    back_ptr= 0
    step= 0
    WebDriverWait(driver,30,0.5).until(EC.presence_of_element_located((By.CLASS_NAME,"x1i10hfl.xjbqb8w")))
    try:
        while flag!= 1:
            flag= 1
            step= 0
            
            more= Web.find_elements(By.CLASS_NAME,"x1i10hfl.xjbqb8w")
            #print(len(more))
            back_ptr= len(more)-1
            while back_ptr>= 0:
                if step> 15:
                    break
                #print(back_ptr)
                
                try:
                    if "檢視另" in more[back_ptr].text or "顯示更多" in more[back_ptr].text or "查看更多" in more[back_ptr].text:
                        driver.execute_script("arguments[0].click();",more[back_ptr])
                        flag= 0
                        #print(more[back_ptr].text)
                        break
                    back_ptr-= 1
                    step+= 1
                except:
                    #print("wrong message")
                    flag= 0
                    break
    except:
        print("your computer is too low")

        #more.clear
    #抓取留言內容
    WebDriverWait(driver,30,0.5).until(EC.presence_of_element_located((By.CSS_SELECTOR,"div.x1y1aw1k.xn6708d > div")))
    comment_div= Web.find_elements(By.CSS_SELECTOR,"div.x1y1aw1k.xn6708d > div")

    print("相關留言:",len(comment_div),"筆\n")

    print("垃圾、標記式留言過濾中......\n")
    for i in range(0,len(comment_div)):
        try:
            comment_div[i].find_element(By.CSS_SELECTOR,"a")
        except:
            comment.append(comment_div[i].text)
    

    #過濾標記式留言
    '''
    for i in range(0,len(comment_div)):
        try:
            comment_div[i].find_element(By.CSS_SELECTOR,"a")
        except:
            if operator.not_("顯示更多" in comment_div[i].text or "\n" in comment_div[i].text):
                comment.append(comment_div[i].text)
    '''

    print("有效樣本留言:",len(comment),"筆\n")
    print("篩選關鍵字及排除非法輸入......\n")

    xlsx_row= 1
    xlsx_col= 1
    for i in range(0,len(comment)):
        if comment[i]!= "作者"and operator.not_("留言……" in comment[i]) and operator.not_("回覆" in comment[i] and "......" in comment[i]) and operator.not_("顯示更多" in comment[i]) and operator.not_("頭號粉絲" in comment[i]):
            if "\n" in comment[i]:
                tmp_comment= comment[i].split('\n')
                sum_comment= ''
                for ptr in range(0,len(tmp_comment)):
                    sum_comment+= tmp_comment[ptr]
                    sum_comment+= ','
                comment[i]= sum_comment

            for k in range(0,len(Keyword)):
                if Keyword[k] in comment[i]:
                    Comment_List.append(comment[i])
                    workpage.cell(xlsx_row,xlsx_col).value= comment[i]
                    xlsx_row+= 1
                    if xlsx_row> 15:
                        xlsx_col+= 1
                        xlsx_row= 1
                    
                    #print(comment[i])
                    break
    print("獲取目標留言:",len(Comment_List),"筆\n")
    print("完成!!!")                
#  貼入目標網站
def data_login(driver, web_account, web_password):

    try:
        WebDriverWait(driver,180,1).until(EC.presence_of_element_located((By.ID,"inputUsername")))
        account= driver.find_element(By.ID,"inputUsername")
        password= driver.find_element(By.ID,"inputPassword")
        button= driver.find_element(By.ID,"signinSubmit")

        while driver.find_element(By.ID,"inputUsername").get_attribute("value")!= web_account:
            account.send_keys(web_account)
        while driver.find_element(By.ID,"inputPassword").get_attribute("value")!= web_password:
            password.send_keys(web_password)

        driver.execute_script("arguments[0].click();",button)
    except:
        print("data_login error")

def data_scan(alt,driver):
    try:
        WebDriverWait(driver,20,0.5).until(EC.presence_of_element_located((By.ID,"inputLink")))
        alt.link= driver.find_element(By.ID,"inputLink").get_attribute("value")
        alt.select_att= driver.find_element(By.ID,"attribute").get_attribute("value")
        alt.select_vic= driver.find_element(By.ID,"viceattribute").get_attribute("value")
        '''select_vic.select_by_value("民生經濟")
        print("select_vic=",select_vic)'''

        #driver.execute_script("arguments[0].value = '你猜一下';", search_button)
        #js = 'arguments[0].removeAttribute("readonly");'
        alt.media_date= driver.find_element(By.ID,"found_date").get_attribute("value")
        '''js = 'arguments[0].removeAttribute("readonly");'
        driver.execute_script(js,media_date)
        driver.execute_script("arguments[0].value = '2021-02-20';", media_date)'''
        alt.china_media= driver.find_element(By.ID,"found_type").get_attribute("value")
        alt.media_name= driver.find_element(By.ID,"found_group").get_attribute("value")
        alt.media_account= driver.find_element(By.ID,"found_account").get_attribute("value")


        alt.source_date= driver.find_element(By.ID, "source_date").get_attribute("value")
        '''js = 'arguments[0].removeAttribute("readonly");'
        driver.execute_script(js,source_date)
        driver.execute_script("arguments[0].value = '2021-02-20';", source_date)'''
        alt.source_media= driver.find_element(By.ID,"source_type").get_attribute("value")
        alt.source_name= driver.find_element(By.ID,"source_group").get_attribute("value")
        alt.source_account= driver.find_element(By.ID,"source_account").get_attribute("value")
        alt.source_link= driver.find_element(By.ID,"source_link").get_attribute("value")

        alt.company_main= driver.find_element(By.ID,"firstgovernment").get_attribute("value")
        alt.company_vice= driver.find_element(By.ID,"secondgovernment").get_attribute("value")
    except:
        print("data_scan error")

def data_write(alt,driver):
    js = 'arguments[0].removeAttribute("readonly");'
    try:
        WebDriverWait(driver,120,2).until(EC.presence_of_element_located((By.ID,"inputTitle")))

        title_string= ""

        if len(Comment_List)== 0:
            return False
        
        content= driver.find_element(By.ID,"inputContent")
        title= driver.find_element(By.ID,"inputTitle")

        num= random.randint(0,len(Comment_List)-1)
        content.send_keys(Comment_List[num])
        title_string+= Comment_List[num]
        title_string+= " "
        Comment_List.pop(num)
        
        driver.execute_script("arguments[0].value = '"+ title_string+ "';", title)

        link= driver.find_element(By.ID,"inputLink")
        driver.execute_script("arguments[0].value = '"+ alt.link+ "';", link)

        select_att= Select(driver.find_element(By.ID,"attribute"))
        try:
            select_att.select_by_value(alt.select_att)
        except:
            print("no input")
        select_vic= Select(driver.find_element(By.ID,"viceattribute"))
        try:
            select_vic.select_by_value(alt.select_vic)
        except:
            print("no input")

        #driver.execute_script("arguments[0].value = '你猜一下';", search_button)
        #js = 'arguments[0].removeAttribute("readonly");'
        media_date= driver.find_element(By.ID,"found_date")
        driver.execute_script(js,media_date)
        driver.execute_script("arguments[0].value = '"+ alt.media_date+ "';", media_date)

        china_media= Select(driver.find_element(By.ID,"found_type"))
        try:
            china_media.select_by_value(alt.china_media)
        except:
            print("no input")

        media_name= driver.find_element(By.ID,"found_group")
        driver.execute_script("arguments[0].value = '"+ alt.media_name+ "';", media_name)
        media_account= driver.find_element(By.ID,"found_account")
        driver.execute_script("arguments[0].value = '"+ alt.media_account+ "';", media_account)
    except:
        print("data_write error 電腦速度不足!!")


    source_date= driver.find_element(By.ID, "source_date")
    driver.execute_script(js,source_date)
    driver.execute_script("arguments[0].value = '"+ alt.source_date+ "';", source_date)

    source_media= Select(driver.find_element(By.ID,"source_type"))
    try:
        source_media.select_by_value(alt.source_media)
    except:
        print("no input")

    source_name= driver.find_element(By.ID,"source_group")
    driver.execute_script("arguments[0].value = '"+ alt.source_name+ "';", source_name)
    source_account= driver.find_element(By.ID,"source_account")
    driver.execute_script("arguments[0].value = '"+ alt.source_account+ "';", source_account)
    source_link= driver.find_element(By.ID,"source_link")
    driver.execute_script("arguments[0].value = '"+ alt.source_link+ "';", source_link)


    company_main= Select(driver.find_element(By.ID,"firstgovernment"))
    try:
        company_main.select_by_value(alt.company_main)
    except:
        print("no input")
    company_vice= Select(driver.find_element(By.ID,"secondgovernment"))
    try:
        company_vice.select_by_value(alt.company_vice)
    except:
        print("no input")
    

    complete= driver.find_element(By.CSS_SELECTOR, "#fmForm > button")
    driver.execute_script("arguments[0].click();", complete)

    return True


#介面---------------------------------

class Window(object):

    def __init__(self):

        self.root = tk.Tk()
        self.root.title("爬蟲")
        self.root.geometry('500x550')
        # Entry
        self.input_account = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_password = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_url = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_keyword = tk.Entry(self.root, width=40, font=('Courier',9))
        self.input_word = tk.Entry(self.root, width=40,  font=('Courier',9))
        self.input_web_account = tk.Entry(self.root,  width= 40, font=('Courier',9))
        self.input_web_password = tk.Entry(self.root, width= 40, font=('Courier',9))
        # Label
        self.label_account = tk.Label(self.root, text="FB帳號: ", font=('Courier',9))
        self.label_password = tk.Label(self.root, text="FB密碼: ", font=('Courier',9))
        self.label_url = tk.Label(self.root, text="請輸入網址: ", font=('Courier',9))    
        self.label_keyword = tk.Label(self.root, text="請輸入關鍵字: ", font=('Courier',9))
        self.label_web_account = tk.Label(self.root, text="目標網站帳號: ", font=('Courier',9))
        self.label_web_password = tk.Label(self.root, text="目標網站密碼: ", font=('Courier',9))
        self.label_word = tk.Label(self.root, text= "手動新增留言",  font=('Courier',9))
        # Button
        self.start_botton = tk.Button(text = "開始",  command=self.start_event, width=30)
        self.clear_botton = tk.Button(text = "清除",  command=self.clear_event, width=30)
        self.scan_botton = tk.Button(text = "掃描",  command=self.scan_event, width=30)
        self.return_botton = tk.Button(text = "送出",  command=self.return_event, width=30)
        self.plus_button= tk.Button(text = "新增",  command=self.plus_event, width=30)
    def gui_arrange(self):
        # Entry
        self.input_account.place(x=110, y=110, height=25)
        self.input_password.place(x=110, y=140, height=25)
        self.input_url.place(x=110, y=170, height=25)  
        self.input_web_account.place(x=110, y=200, height=25)
        self.input_web_password.place(x=110, y=230, height=25)
        self.input_keyword.place(x=110, y=260, height=25)
        self.input_word.place(x=110, y=290, height=25)
        # Label
        self.label_account.place(x=30, y=110)
        self.label_password.place(x=30, y=140)
        self.label_url.place(x=25, y= 170) 
        self.label_web_account.place(x=15, y=200)
        self.label_web_password.place(x=15, y=230)
        self.label_keyword.place(x=15, y= 260)
        self.label_word.place(x=15,y=290)
        # Botton
        self.start_botton.place(x=140, y=350, height=40,  width = 220)
        self.clear_botton.place(x=140, y=400, height=30,  width = 220)
        self.scan_botton.place(x=5, y=350, height=40,  width = 120)
        self.return_botton.place(x=140, y=440, height=30,  width = 220)
        self.plus_button.place(x=375, y=350, height=40,  width = 120)
        # Web Element
        self.driver=  uc.Chrome(use_subprocess= True, chrome_options= options)

        self.link= ''
        self.select_att= '請選擇'
        self.select_vic= '請選擇'
        self.media_date= ''
        self.media_name= ''
        self.media_account= ''
        self.china_media= '臉書個人專頁'
        self.source_media= '臉書個人專頁'
        self.source_date= ''
        self.source_name= ''
        self.source_account= ''
        self.source_link= ''
        self.company_main= '請選擇'
        self.company_vice= '無'

    def clear_event(self):
        self.input_url.delete(0, 'end')
        self.input_keyword.delete(0, 'end')
        self.input_account.delete(0, 'end')
        self.input_password.delete(0, 'end')
        self.input_web_account.delete(0, 'end')
        self.input_web_password.delete(0, 'end')
        self.input_word.delete(0, 'end')

    def start_event(self):
        url = self.input_url.get()
        keyword = self.input_keyword.get()
        account = self.input_account.get()
        password = self.input_password.get()
        web_account= self.input_web_account.get()
        web_password= self.input_web_password.get()

        self.clear_event()
        Comment_List.clear()
        comment.clear()
        print("清除暫存留言")

        self.driver.get(url)
        scrape(self.driver, account, password, keyword)
        try:
            workbook.save("./new.xlsx")
        except:
            print("檔案可能開啟，造成錯誤")

        time.sleep(5)
        self.driver.get("https://btsp.servehttp.com/users/signin")
        data_login(self.driver, web_account, web_password)
        #data_write(self,self.driver)
        # 接口


    def plus_event(self):
        Word= self.input_word.get()
        tmp_Word= Word.split(' ')
        self.input_word.delete(0, 'end')
        for i in range(0, len(tmp_Word)):
            if tmp_Word[i]!= '':
                Comment_List.append(tmp_Word[i])
                print("成功新增 ->",tmp_Word[i],"<-")

    def scan_event(self):
        data_scan(self,self.driver)

    def return_event(self):
        while data_write(self,self.driver):
            try:
                WebDriverWait(self.driver,180,1).until(EC.presence_of_element_located((By.CLASS_NAME,"dt-center")))
                print("傳送完成!!")
            except:


                print("傳送失敗!!")

            self.driver.get("https://btsp.servehttp.com/report/report")
        print("全部傳送完畢!")

def main():
    window = Window()
    window.gui_arrange()
    tk.mainloop()

if __name__ == '__main__':  
    main()



#--------------------------------





